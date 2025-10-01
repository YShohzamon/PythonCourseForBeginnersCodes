import openpyxl as xl # excel biln ishlash uchun.
from openpyxl.chart import BarChart, Reference

def prosses_workbook():
    wb = xl.load_workbook("information.xlsx")
    sheet = wb["Sheet1"]
    # cell = sheet["a1"]
    cell = sheet.cell(1 , 1)
    # print(cell.value)
    # print(sheet.max_row)

    for row in range(2, sheet.max_row+1):
        # print(row)
        cell = sheet.cell(row, 3)
        corrented_cell = cell.value*0.9
        corrented_cell_price = sheet.cell(row, 4)
        corrented_cell_price.value = corrented_cell

    values =  Reference(sheet, min_row=2, max_row=sheet.max_row , min_col=4 , max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save("information2.xlsx")
